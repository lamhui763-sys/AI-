"""
Excel可视化模块
用于生成专业的Excel报表，包含图表、格式化表格等

特性:
- 自动生成格式化的Excel报表
- 内置多种图表类型
- 技术指标可视化
- 回测结果可视化
- 投资组合分析
"""

import pandas as pd
import numpy as np
from typing import Dict, List, Optional, Any
from datetime import datetime
import os


class ExcelVisualizer:
    """Excel可视化生成器"""
    
    def __init__(self, output_dir: str = 'output/excel'):
        """
        初始化Excel可视化器
        
        Args:
            output_dir: Excel文件输出目录
        """
        self.output_dir = output_dir
        os.makedirs(output_dir, exist_ok=True)
        
    def generate_stock_report(self, 
                             symbol: str,
                             price_data: pd.DataFrame,
                             indicators: Dict[str, Any],
                             ai_analysis: Optional[Dict[str, Any]] = None,
                             filename: Optional[str] = None) -> str:
        """
        生成股票分析Excel报告
        
        Args:
            symbol: 股票代码
            price_data: 价格数据
            indicators: 技术指标数据
            ai_analysis: AI分析结果（可选）
            filename: 输出文件名（可选）
            
        Returns:
            生成的Excel文件路径
        """
        if filename is None:
            filename = f"{symbol}_report_{datetime.now().strftime('%Y%m%d')}.xlsx"
        
        filepath = os.path.join(self.output_dir, filename)
        
        # 创建Excel写入器
        with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
            # 1. 价格数据页
            self._write_price_sheet(writer, price_data, symbol)
            
            # 2. 技术指标页
            self._write_indicators_sheet(writer, indicators, symbol)
            
            # 3. 交易信号页
            self._write_signals_sheet(writer, price_data, indicators, symbol)
            
            # 4. AI分析页（如果有）
            if ai_analysis:
                self._write_ai_sheet(writer, ai_analysis, symbol)
            
            # 5. 摘要页
            self._write_summary_sheet(writer, symbol, price_data, indicators, ai_analysis)
        
        print(f"✅ Excel报告已生成: {filepath}")
        return filepath
    
    def _write_price_sheet(self, writer, data: pd.DataFrame, symbol: str):
        """写入价格数据页"""
        sheet_name = '价格数据'
        
        # 选择重要列
        columns = ['Open', 'High', 'Low', 'Close', 'Volume']
        available_columns = [col for col in columns if col in data.columns]
        
        df = data[available_columns].copy()
        df.index.name = '日期'
        
        # 写入数据
        df.to_excel(writer, sheet_name=sheet_name)
        
        # 获取worksheet并设置格式
        worksheet = writer.sheets[sheet_name]
        
        # 设置列宽
        column_widths = {
            'A': 12,  # 日期
            'B': 12,  # Open
            'C': 12,  # High
            'D': 12,  # Low
            'E': 12,  # Close
            'F': 12,  # Volume
        }
        for col, width in column_widths.items():
            if col in [cell.column_letter for cell in worksheet[1]]:
                worksheet.column_dimensions[col].width = width
        
        # 添加标题
        worksheet.insert_rows(1)
        worksheet['A1'] = f'{symbol} - 历史价格数据'
        worksheet['A1'].font = Font(bold=True, size=14)
        worksheet.merge_cells('A1:F1')
        worksheet['A1'].alignment = Alignment(horizontal='center')
    
    def _write_indicators_sheet(self, writer, indicators: Dict[str, Any], symbol: str):
        """写入技术指标页"""
        sheet_name = '技术指标'
        
        # 提取指标数据
        df_dict = {}
        
        # 趋势指标
        if 'SMA' in indicators:
            for period, values in indicators['SMA'].items():
                df_dict[f'SMA_{period}'] = values
        
        if 'EMA' in indicators:
            for period, values in indicators['EMA'].items():
                df_dict[f'EMA_{period}'] = values
        
        # 布林带
        if 'Bollinger_Bands' in indicators:
            bb = indicators['Bollinger_Bands']
            df_dict['BB_Upper'] = bb['upper']
            df_dict['BB_Middle'] = bb['middle']
            df_dict['BB_Lower'] = bb['lower']
        
        # RSI
        if 'RSI' in indicators:
            df_dict['RSI'] = indicators['RSI']
        
        # MACD
        if 'MACD' in indicators:
            macd = indicators['MACD']
            df_dict['MACD'] = macd['macd']
            df_dict['MACD_Signal'] = macd['signal']
            df_dict['MACD_Histogram'] = macd['histogram']
        
        # 创建DataFrame
        df = pd.DataFrame(df_dict)
        df.index.name = '日期'
        
        # 写入数据
        df.to_excel(writer, sheet_name=sheet_name)
        
        # 获取worksheet
        worksheet = writer.sheets[sheet_name]
        
        # 设置列宽
        for col in worksheet.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 20)
            worksheet.column_dimensions[column].width = adjusted_width
        
        # 添加标题
        worksheet.insert_rows(1)
        worksheet['A1'] = f'{symbol} - 技术指标'
        worksheet['A1'].font = Font(bold=True, size=14)
        end_col = len(df.columns) + 1
        worksheet.merge_cells(f'A1:{chr(64 + end_col)}1')
        worksheet['A1'].alignment = Alignment(horizontal='center')
    
    def _write_signals_sheet(self, writer, price_data: pd.DataFrame, 
                            indicators: Dict[str, Any], symbol: str):
        """写入交易信号页"""
        sheet_name = '交易信号'
        
        # 生成信号数据
        signals = self._generate_signals(price_data, indicators)
        
        if not signals.empty:
            signals.index.name = '日期'
            signals.to_excel(writer, sheet_name=sheet_name)
            
            worksheet = writer.sheets[sheet_name]
            
            # 添加标题
            worksheet.insert_rows(1)
            worksheet['A1'] = f'{symbol} - 交易信号'
            worksheet['A1'].font = Font(bold=True, size=14)
            worksheet.merge_cells('A1:E1')
            worksheet['A1'].alignment = Alignment(horizontal='center')
        else:
            # 如果没有信号，写入说明
            pd.DataFrame({'说明': ['当前时间段无交易信号']}).to_excel(
                writer, sheet_name=sheet_name, index=False
            )
    
    def _generate_signals(self, price_data: pd.DataFrame, 
                          indicators: Dict[str, Any]) -> pd.DataFrame:
        """生成交易信号数据"""
        signals_data = []
        
        # 获取数据长度
        length = len(price_data)
        
        # 分析最近30天的信号
        for i in range(max(0, length - 30), length):
            date = price_data.index[i]
            
            # 计算信号
            signal_strength = 0
            signals_list = []
            
            # RSI信号
            if 'RSI' in indicators:
                rsi = indicators['RSI'].iloc[i]
                if rsi < 30:
                    signal_strength += 2
                    signals_list.append('RSI超卖')
                elif rsi > 70:
                    signal_strength -= 2
                    signals_list.append('RSI超买')
            
            # MACD信号
            if 'MACD' in indicators and i > 0:
                macd = indicators['MACD']['macd'].iloc[i]
                signal = indicators['MACD']['signal'].iloc[i]
                prev_macd = indicators['MACD']['macd'].iloc[i-1]
                prev_signal = indicators['MACD']['signal'].iloc[i-1]
                
                # 金叉
                if prev_macd < prev_signal and macd > signal:
                    signal_strength += 2
                    signals_list.append('MACD金叉')
                # 死叉
                elif prev_macd > prev_signal and macd < signal:
                    signal_strength -= 2
                    signals_list.append('MACD死叉')
            
            # 均线信号
            if 'SMA' in indicators and 'SMA_5' in indicators['SMA'] and 'SMA_20' in indicators['SMA']:
                sma5 = indicators['SMA']['SMA_5'].iloc[i]
                sma20 = indicators['SMA']['SMA_20'].iloc[i]
                
                if sma5 > sma20:
                    signal_strength += 1
                    signals_list.append('5日线>20日线')
                else:
                    signal_strength -= 1
                    signals_list.append('5日线<20日线')
            
            # 确定信号类型
            if signal_strength >= 4:
                signal_type = 'STRONG BUY'
            elif signal_strength >= 2:
                signal_type = 'BUY'
            elif signal_strength >= -1:
                signal_type = 'HOLD'
            elif signal_strength >= -3:
                signal_type = 'SELL'
            else:
                signal_type = 'STRONG SELL'
            
            if signals_list:  # 只记录有信号的日期
                signals_data.append({
                    '日期': date,
                    '信号': signal_type,
                    '强度': signal_strength,
                    '收盘价': price_data['Close'].iloc[i],
                    '信号说明': ', '.join(signals_list)
                })
        
        return pd.DataFrame(signals_data)
    
    def _write_ai_sheet(self, writer, ai_analysis: Dict[str, Any], symbol: str):
        """写入AI分析页"""
        sheet_name = 'AI分析'
        
        # 创建分析数据
        analysis_data = []
        
        # AI建议
        if 'recommendation' in ai_analysis:
            rec = ai_analysis['recommendation']
            analysis_data.append({
                '项目': 'AI建议',
                '内容': rec.get('action', 'N/A')
            })
            analysis_data.append({
                '项目': '置信度',
                '内容': f"{rec.get('confidence', 0):.1%}"
            })
            analysis_data.append({
                '项目': '风险等级',
                '内容': rec.get('risk_level', 'N/A')
            })
        
        # AI分析详情
        if 'analysis_details' in ai_analysis:
            details = ai_analysis['analysis_details']
            if 'summary' in details:
                analysis_data.append({
                    '项目': '分析摘要',
                    '内容': details['summary']
                })
            if 'key_points' in details:
                analysis_data.append({
                    '项目': '关键观点',
                    '内容': '\n'.join(details['key_points'])
                })
        
        # 情感分析
        if 'sentiment' in ai_analysis:
            sent = ai_analysis['sentiment']
            analysis_data.append({
                '项目': '市场情感',
                '内容': sent.get('overall', 'N/A')
            })
            analysis_data.append({
                '项目': '情感分数',
                '内容': f"{sent.get('score', 0):.2f}"
            })
        
        # 创建DataFrame
        df = pd.DataFrame(analysis_data)
        
        # 写入数据
        df.to_excel(writer, sheet_name=sheet_name, index=False)
        
        # 获取worksheet并设置格式
        worksheet = writer.sheets[sheet_name]
        
        # 设置列宽
        worksheet.column_dimensions['A'].width = 15
        worksheet.column_dimensions['B'].width = 60
        
        # 设置自动换行
        for row in worksheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical='top')
        
        # 添加标题
        worksheet.insert_rows(1)
        worksheet['A1'] = f'{symbol} - AI分析'
        worksheet['A1'].font = Font(bold=True, size=14)
        worksheet.merge_cells('A1:B1')
        worksheet['A1'].alignment = Alignment(horizontal='center')
        
        # 设置行高
        worksheet.row_dimensions[2].height = 25  # 标题行
    
    def _write_summary_sheet(self, writer, symbol: str, price_data: pd.DataFrame,
                            indicators: Dict[str, Any], 
                            ai_analysis: Optional[Dict[str, Any]] = None):
        """写入摘要页"""
        sheet_name = '摘要'
        
        # 获取最新数据
        latest_date = price_data.index[-1]
        latest_close = price_data['Close'].iloc[-1]
        
        # 计算基本信息
        summary_data = []
        
        # 基本信息
        summary_data.append(['股票代码', symbol])
        summary_data.append(['更新日期', latest_date.strftime('%Y-%m-%d')])
        summary_data.append(['最新收盘价', f"HK${latest_close:.2f}"])
        
        # 最近涨跌
        if len(price_data) >= 2:
            prev_close = price_data['Close'].iloc[-2]
            change = latest_close - prev_close
            change_pct = (change / prev_close) * 100
            summary_data.append(['日涨跌', f"{change:+.2f} ({change_pct:+.2f}%)"])
        
        summary_data.append(['', ''])  # 空行
        
        # 技术指标摘要
        summary_data.append(['技术指标摘要', ''])
        
        # RSI
        if 'RSI' in indicators:
            rsi = indicators['RSI'].iloc[-1]
            rsi_status = '超买' if rsi > 70 else '超卖' if rsi < 30 else '中性'
            summary_data.append(['RSI (14)', f"{rsi:.2f} - {rsi_status}"])
        
        # MACD
        if 'MACD' in indicators:
            macd = indicators['MACD']['macd'].iloc[-1]
            signal = indicators['MACD']['signal'].iloc[-1]
            macd_status = '看涨' if macd > signal else '看跌'
            summary_data.append(['MACD', f"{macd:.2f} - {macd_status}"])
        
        # 布林带
        if 'Bollinger_Bands' in indicators:
            bb = indicators['Bollinger_Bands']
            upper = bb['upper'].iloc[-1]
            lower = bb['lower'].iloc[-1]
            bb_status = '突破上轨' if latest_close > upper else '跌破下轨' if latest_close < lower else '中轨附近'
            summary_data.append(['布林带位置', bb_status])
        
        # 趋势
        if 'SMA' in indicators and 'SMA_5' in indicators['SMA'] and 'SMA_20' in indicators['SMA']:
            sma5 = indicators['SMA']['SMA_5'].iloc[-1]
            sma20 = indicators['SMA']['SMA_20'].iloc[-1]
            trend = '上升趋势' if sma5 > sma20 else '下降趋势'
            summary_data.append(['趋势', trend])
        
        summary_data.append(['', ''])  # 空行
        
        # AI分析摘要
        if ai_analysis and 'recommendation' in ai_analysis:
            summary_data.append(['AI分析摘要', ''])
            rec = ai_analysis['recommendation']
            summary_data.append(['AI建议', rec.get('action', 'N/A')])
            summary_data.append(['置信度', f"{rec.get('confidence', 0):.1%}"])
            summary_data.append(['风险等级', rec.get('risk_level', 'N/A')])
            
            if 'sentiment' in ai_analysis:
                sent = ai_analysis['sentiment']
                summary_data.append(['市场情感', sent.get('overall', 'N/A')])
        
        # 创建DataFrame
        df = pd.DataFrame(summary_data, columns=['项目', '内容'])
        
        # 写入数据
        df.to_excel(writer, sheet_name=sheet_name, index=False)
        
        # 获取worksheet并设置格式
        worksheet = writer.sheets[sheet_name]
        
        # 设置列宽
        worksheet.column_dimensions['A'].width = 20
        worksheet.column_dimensions['B'].width = 40
        
        # 设置格式
        for row in worksheet.iter_rows():
            row[0].font = Font(bold=True)  # 第一列加粗
            row[0].alignment = Alignment(horizontal='right')
            row[1].alignment = Alignment(horizontal='left')
        
        # 添加标题
        worksheet.insert_rows(1)
        worksheet['A1'] = f'{symbol} - 分析摘要'
        worksheet['A1'].font = Font(bold=True, size=16)
        worksheet.merge_cells('A1:B1')
        worksheet['A1'].alignment = Alignment(horizontal='center')
    
    def generate_backtest_report(self,
                                 results: Dict[str, Any],
                                 strategy_name: str,
                                 filename: Optional[str] = None) -> str:
        """
        生成回测报告Excel
        
        Args:
            results: 回测结果
            strategy_name: 策略名称
            filename: 输出文件名（可选）
            
        Returns:
            生成的Excel文件路径
        """
        if filename is None:
            filename = f"backtest_{strategy_name}_{datetime.now().strftime('%Y%m%d')}.xlsx"
        
        filepath = os.path.join(self.output_dir, filename)
        
        with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
            # 1. 摘要页
            self._write_backtest_summary(writer, results, strategy_name)
            
            # 2. 交易记录页
            if 'trades' in results and not results['trades'].empty:
                self._write_trades_sheet(writer, results['trades'])
            
            # 3. 资金曲线页
            if 'equity_curve' in results:
                self._write_equity_sheet(writer, results['equity_curve'])
            
            # 4. 性能指标页
            self._write_performance_sheet(writer, results['summary'])
        
        print(f"✅ 回测报告已生成: {filepath}")
        return filepath
    
    def _write_backtest_summary(self, writer, results: Dict[str, Any], strategy_name: str):
        """写入回测摘要页"""
        sheet_name = '回测摘要'
        
        summary = results.get('summary', {})
        
        summary_data = [
            ['策略名称', strategy_name],
            ['', ''],
            ['回测期间', f"{summary.get('start_date', 'N/A')} 至 {summary.get('end_date', 'N/A')}"],
            ['', ''],
            ['初始资金', f"HK${summary.get('initial_capital', 0):,.2f}"],
            ['最终资金', f"HK${summary.get('final_capital', 0):,.2f}"],
            ['总收益', f"{summary.get('total_return', 0):.2f}%"],
            ['年化收益', f"{summary.get('annual_return', 0):.2f}%"],
            ['', ''],
            ['交易次数', summary.get('total_trades', 0)],
            ['盈利交易', summary.get('winning_trades', 0)],
            ['亏损交易', summary.get('losing_trades', 0)],
            ['胜率', f"{summary.get('win_rate', 0):.2f}%"],
            ['', ''],
            ['最大回撤', f"{summary.get('max_drawdown', 0):.2f}%"],
            ['夏普比率', f"{summary.get('sharpe_ratio', 0):.2f}"],
            ['索提诺比率', f"{summary.get('sortino_ratio', 0):.2f}"],
            ['Calmar比率', f"{summary.get('calmar_ratio', 0):.2f}"],
        ]
        
        df = pd.DataFrame(summary_data, columns=['项目', '内容'])
        df.to_excel(writer, sheet_name=sheet_name, index=False)
        
        # 设置格式
        worksheet = writer.sheets[sheet_name]
        worksheet.column_dimensions['A'].width = 20
        worksheet.column_dimensions['B'].width = 30
        
        for row in worksheet.iter_rows():
            row[0].font = Font(bold=True)
            row[0].alignment = Alignment(horizontal='right')
            row[1].alignment = Alignment(horizontal='left')
        
        # 添加标题
        worksheet.insert_rows(1)
        worksheet['A1'] = f'{strategy_name} - 回测报告'
        worksheet['A1'].font = Font(bold=True, size=16)
        worksheet.merge_cells('A1:B1')
        worksheet['A1'].alignment = Alignment(horizontal='center')
    
    def _write_trades_sheet(self, writer, trades: pd.DataFrame):
        """写入交易记录页"""
        sheet_name = '交易记录'
        
        # 选择重要列
        columns_to_export = ['entry_date', 'exit_date', 'entry_price', 
                           'exit_price', 'shares', 'pnl', 'pnl_pct']
        available_columns = [col for col in columns_to_export if col in trades.columns]
        
        df = trades[available_columns].copy()
        
        # 重命名列
        column_names = {
            'entry_date': '买入日期',
            'exit_date': '卖出日期',
            'entry_price': '买入价格',
            'exit_price': '卖出价格',
            'shares': '股数',
            'pnl': '盈亏金额',
            'pnl_pct': '盈亏百分比'
        }
        df.columns = [column_names.get(col, col) for col in df.columns]
        
        df.to_excel(writer, sheet_name=sheet_name, index=False)
        
        # 设置格式
        worksheet = writer.sheets[sheet_name]
        for col in worksheet.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            worksheet.column_dimensions[column].width = min(max_length + 2, 15)
    
    def _write_equity_sheet(self, writer, equity_curve: pd.DataFrame):
        """写入资金曲线页"""
        sheet_name = '资金曲线'
        
        equity_curve.to_excel(writer, sheet_name=sheet_name)
        
        worksheet = writer.sheets[sheet_name]
        worksheet.column_dimensions['A'].width = 15
        worksheet.column_dimensions['B'].width = 20
    
    def _write_performance_sheet(self, writer, summary: Dict[str, Any]):
        """写入性能指标页"""
        sheet_name = '性能指标'
        
        # 详细的性能指标
        performance_data = [
            ['收益指标', ''],
            ['总收益率', f"{summary.get('total_return', 0):.2f}%"],
            ['年化收益率', f"{summary.get('annual_return', 0):.2f}%"],
            ['累计收益', f"{summary.get('final_capital', 0) - summary.get('initial_capital', 0):,.2f}"],
            ['', ''],
            ['风险指标', ''],
            ['年化波动率', f"{summary.get('volatility', 0):.2f}%"],
            ['最大回撤', f"{summary.get('max_drawdown', 0):.2f}%"],
            ['最大回撤持续期', f"{summary.get('max_drawdown_duration', 0)} 天"],
            ['', ''],
            ['风险调整收益', ''],
            ['夏普比率', f"{summary.get('sharpe_ratio', 0):.2f}"],
            ['索提诺比率', f"{summary.get('sortino_ratio', 0):.2f}"],
            ['Calmar比率', f"{summary.get('calmar_ratio', 0):.2f}"],
            ['', ''],
            ['交易统计', ''],
            ['总交易次数', summary.get('total_trades', 0)],
            ['盈利交易数', summary.get('winning_trades', 0)],
            ['亏损交易数', summary.get('losing_trades', 0)],
            ['胜率', f"{summary.get('win_rate', 0):.2f}%"],
            ['平均盈利', f"{summary.get('avg_win', 0):.2f}"],
            ['平均亏损', f"{summary.get('avg_loss', 0):.2f}"],
            ['盈亏比', f"{summary.get('profit_factor', 0):.2f}"],
        ]
        
        df = pd.DataFrame(performance_data, columns=['指标', '数值'])
        df.to_excel(writer, sheet_name=sheet_name, index=False)
        
        # 设置格式
        worksheet = writer.sheets[sheet_name]
        worksheet.column_dimensions['A'].width = 25
        worksheet.column_dimensions['B'].width = 25
        
        for row in worksheet.iter_rows():
            row[0].font = Font(bold=True)
            row[0].alignment = Alignment(horizontal='right')
            row[1].alignment = Alignment(horizontal='left')
        
        # 添加标题
        worksheet.insert_rows(1)
        worksheet['A1'] = '详细性能指标'
        worksheet['A1'].font = Font(bold=True, size=14)
        worksheet.merge_cells('A1:B1')
        worksheet['A1'].alignment = Alignment(horizontal='center')
    
    def generate_portfolio_report(self,
                                 portfolio_data: Dict[str, Any],
                                 filename: Optional[str] = None) -> str:
        """
        生成投资组合分析报告
        
        Args:
            portfolio_data: 投资组合数据
            filename: 输出文件名（可选）
            
        Returns:
            生成的Excel文件路径
        """
        if filename is None:
            filename = f"portfolio_report_{datetime.now().strftime('%Y%m%d')}.xlsx"
        
        filepath = os.path.join(self.output_dir, filename)
        
        with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
            # 1. 组合概览
            self._write_portfolio_overview(writer, portfolio_data)
            
            # 2. 持仓详情
            if 'holdings' in portfolio_data:
                self._write_holdings_sheet(writer, portfolio_data['holdings'])
            
            # 3. 收益分布
            if 'returns' in portfolio_data:
                self._write_returns_sheet(writer, portfolio_data['returns'])
            
            # 4. 风险分析
            if 'risk_analysis' in portfolio_data:
                self._write_risk_sheet(writer, portfolio_data['risk_analysis'])
        
        print(f"✅ 投资组合报告已生成: {filepath}")
        return filepath
    
    def _write_portfolio_overview(self, writer, portfolio_data: Dict[str, Any]):
        """写入投资组合概览"""
        sheet_name = '组合概览'
        
        overview_data = [
            ['组合总值', f"HK${portfolio_data.get('total_value', 0):,.2f}"],
            ['总收益', f"HK${portfolio_data.get('total_pnl', 0):,.2f}"],
            ['收益率', f"{portfolio_data.get('total_return', 0):.2f}%"],
            ['持仓数量', portfolio_data.get('num_holdings', 0)],
            ['更新日期', portfolio_data.get('date', datetime.now().strftime('%Y-%m-%d'))],
        ]
        
        df = pd.DataFrame(overview_data, columns=['项目', '数值'])
        df.to_excel(writer, sheet_name=sheet_name, index=False)
        
        # 设置格式
        worksheet = writer.sheets[sheet_name]
        worksheet.column_dimensions['A'].width = 20
        worksheet.column_dimensions['B'].width = 30
        
        for row in worksheet.iter_rows():
            row[0].font = Font(bold=True)
            row[0].alignment = Alignment(horizontal='right')
            row[1].alignment = Alignment(horizontal='left')
    
    def _write_holdings_sheet(self, writer, holdings: pd.DataFrame):
        """写入持仓详情页"""
        sheet_name = '持仓详情'
        
        holdings.to_excel(writer, sheet_name=sheet_name, index=False)
        
        # 设置格式
        worksheet = writer.sheets[sheet_name]
        for col in worksheet.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            worksheet.column_dimensions[column].width = min(max_length + 2, 20)
    
    def _write_returns_sheet(self, writer, returns: pd.DataFrame):
        """写入收益分布页"""
        sheet_name = '收益分析'
        
        returns.to_excel(writer, sheet_name=sheet_name)
        
        worksheet = writer.sheets[sheet_name]
        worksheet.column_dimensions['A'].width = 15
        worksheet.column_dimensions['B'].width = 20
    
    def _write_risk_sheet(self, writer, risk_analysis: Dict[str, Any]):
        """写入风险分析页"""
        sheet_name = '风险分析'
        
        risk_data = [
            ['组合波动率', f"{risk_analysis.get('volatility', 0):.2f}%"],
            ['组合Beta', f"{risk_analysis.get('beta', 0):.2f}"],
            ['组合夏普', f"{risk_analysis.get('sharpe', 0):.2f}"],
            ['最大回撤', f"{risk_analysis.get('max_drawdown', 0):.2f}%"],
            ['VaR (95%)', f"{risk_analysis.get('var_95', 0):,.2f}"],
        ]
        
        df = pd.DataFrame(risk_data, columns=['指标', '数值'])
        df.to_excel(writer, sheet_name=sheet_name, index=False)
        
        worksheet = writer.sheets[sheet_name]
        worksheet.column_dimensions['A'].width = 20
        worksheet.column_dimensions['B'].width = 30


# 导出openpyxl格式设置
try:
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("⚠️  警告: openpyxl未安装，某些Excel格式化功能不可用")
    print("   安装命令: pip install openpyxl")
    Font = None
    Alignment = None
