"""
可视化模块使用示例
展示如何使用Excel可视化和Web界面

功能:
- Excel报表生成
- 图表创建
- Web界面使用
"""

import sys
import os

# 添加src到路径
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..'))

from src.ai_inv.data_fetcher import DataFetcher
from src.ai_inv.indicators import TechnicalIndicators
from src.ai_inv.technical_analyzer import TechnicalAnalyzer
from src.ai_inv.excel_visualizer import ExcelVisualizer


def example1_generate_stock_report():
    """示例1: 生成股票分析Excel报告"""
    print("=" * 60)
    print("示例1: 生成股票分析Excel报告")
    print("=" * 60)
    
    # 创建可视化器
    visualizer = ExcelVisualizer(output_dir='output/excel')
    
    # 创建数据获取器
    fetcher = DataFetcher()
    
    # 创建技术分析器
    analyzer = TechnicalAnalyzer()
    
    # 获取数据并分析
    symbol = '^HSI'
    print(f"\n正在分析 {symbol}...")
    
    # 获取价格数据
    data = fetcher.get_historical_data(symbol, period='1y')
    
    # 计算技术指标
    indicators = TechnicalIndicators().calculate_all_indicators(data)
    
    # 获取交易信号
    signal = analyzer.get_trading_signal(symbol)
    
    # 生成Excel报告
    filepath = visualizer.generate_stock_report(
        symbol=symbol,
        price_data=data,
        indicators=indicators,
        ai_analysis=None,  # 可选：添加AI分析
        filename=None  # 自动生成文件名
    )
    
    print(f"\n✅ 报告已生成: {filepath}")
    print("\n报告包含以下工作表:")
    print("  1. 价格数据 - 历史价格和成交量")
    print("  2. 技术指标 - 所有技术指标数据")
    print("  3. 交易信号 - 交易信号记录")
    print("  4. AI分析 - AI分析结果（如果提供）")
    print("  5. 摘要 - 关键信息摘要")


def example2_generate_backtest_report():
    """示例2: 生成回测报告"""
    print("\n" + "=" * 60)
    print("示例2: 生成回测报告")
    print("=" * 60)
    
    from src.ai_inv.backtester import BacktestEngine, MAStrategy
    
    # 创建可视化器
    visualizer = ExcelVisualizer(output_dir='output/excel')
    
    # 创建回测引擎
    engine = BacktestEngine(
        initial_capital=100000,
        commission=0.001,
        slippage=0.0005
    )
    
    # 创建策略
    strategy = MAStrategy(short_period=5, long_period=20)
    
    # 获取数据
    fetcher = DataFetcher()
    data = fetcher.get_historical_data('^HSI', period='2y')
    
    if data is not None and not data.empty:
        print("\n正在运行回测...")
        
        # 运行回测
        results = engine.run_backtest(data, strategy)
        
        # 生成Excel报告
        filepath = visualizer.generate_backtest_report(
            results=results,
            strategy_name='MA_5_20',
            filename=None
        )
        
        print(f"\n✅ 回测报告已生成: {filepath}")
        print("\n报告包含以下工作表:")
        print("  1. 回测摘要 - 关键回测指标")
        print("  2. 交易记录 - 所有交易明细")
        print("  3. 资金曲线 - 资金变化情况")
        print("  4. 性能指标 - 详细性能分析")
        
        # 显示关键指标
        print("\n关键回测指标:")
        summary = results['summary']
        print(f"  总收益率: {summary['total_return']:.2f}%")
        print(f"  年化收益率: {summary['annual_return']:.2f}%")
        print(f"  夏普比率: {summary['sharpe_ratio']:.2f}")
        print(f"  最大回撤: {summary['max_drawdown']:.2f}%")
        print(f"  胜率: {summary['win_rate']:.2f}%")


def example3_generate_portfolio_report():
    """示例3: 生成投资组合报告"""
    print("\n" + "=" * 60)
    print("示例3: 生成投资组合报告")
    print("=" * 60)
    
    import pandas as pd
    import numpy as np
    from datetime import datetime, timedelta
    
    # 创建可视化器
    visualizer = ExcelVisualizer(output_dir='output/excel')
    
    # 创建模拟持仓数据
    print("\n创建模拟投资组合...")
    
    symbols = ['6158.HK', '7200.HK', '^HSI']
    holdings = pd.DataFrame({
        'symbol': symbols,
        'shares': [10000, 5000, 100],
        'avg_price': [0.50, 10.00, 28000.0]
    })
    
    # 获取当前价格
    fetcher = DataFetcher()
    current_prices = {}
    for symbol in symbols:
        data = fetcher.get_historical_data(symbol, period='5d')
        if data is not None and not data.empty:
            current_prices[symbol] = data['Close'].iloc[-1]
        else:
            current_prices[symbol] = holdings[holdings['symbol'] == symbol]['avg_price'].values[0]
    
    # 计算持仓价值
    holdings['current_price'] = holdings['symbol'].map(current_prices)
    holdings['market_value'] = holdings['shares'] * holdings['current_price']
    holdings['cost_basis'] = holdings['shares'] * holdings['avg_price']
    holdings['pnl'] = holdings['market_value'] - holdings['cost_basis']
    holdings['pnl_pct'] = (holdings['pnl'] / holdings['cost_basis']) * 100
    
    # 创建投资组合数据
    portfolio_data = {
        'total_value': holdings['market_value'].sum(),
        'total_pnl': holdings['pnl'].sum(),
        'total_return': (holdings['pnl'].sum() / holdings['cost_basis'].sum()) * 100,
        'num_holdings': len(holdings),
        'date': datetime.now().strftime('%Y-%m-%d'),
        'holdings': holdings,
        'returns': pd.DataFrame({
            '日期': [datetime.now() - timedelta(days=i) for i in range(30, 0, -1)],
            '收益率': np.random.normal(0.001, 0.02, 30)
        }),
        'risk_analysis': {
            'volatility': 15.5,
            'beta': 1.2,
            'sharpe': 1.8,
            'max_drawdown': -8.5,
            'var_95': -5000.0
        }
    }
    
    # 生成投资组合报告
    filepath = visualizer.generate_portfolio_report(
        portfolio_data=portfolio_data,
        filename=None
    )
    
    print(f"\n✅ 投资组合报告已生成: {filepath}")
    print("\n报告包含以下工作表:")
    print("  1. 组合概览 - 投资组合整体信息")
    print("  2. 持仓详情 - 每个持仓的详细信息")
    print("  3. 收益分析 - 收益分布情况")
    print("  4. 风险分析 - 风险指标分析")
    
    # 显示组合摘要
    print("\n投资组合摘要:")
    print(f"  组合总值: HK${portfolio_data['total_value']:,.2f}")
    print(f"  总盈亏: HK${portfolio_data['total_pnl']:,.2f}")
    print(f"  收益率: {portfolio_data['total_return']:.2f}%")
    print(f"  持仓数量: {portfolio_data['num_holdings']}")


def example4_custom_excel_format():
    """示例4: 自定义Excel格式"""
    print("\n" + "=" * 60)
    print("示例4: 自定义Excel格式")
    print("=" * 60)
    
    print("\n提示:")
    print("  - ExcelVisualizer自动处理格式化")
    print("  - 包括字体、颜色、对齐等")
    print("  - 自动调整列宽")
    print("  - 支持合并单元格")
    print("  - 添加标题和摘要")
    
    # 获取数据并生成报告
    fetcher = DataFetcher()
    data = fetcher.get_historical_data('6158.HK', period='6mo')
    
    if data is not None and not data.empty:
        indicators = TechnicalIndicators().calculate_all_indicators(data)
        
        visualizer = ExcelVisualizer(output_dir='output/excel')
        filepath = visualizer.generate_stock_report(
            symbol='6158.HK',
            price_data=data,
            indicators=indicators
        )
        
        print(f"\n✅ 已生成格式化报告: {filepath}")
        print("\n使用Excel打开文件查看格式效果!")


def example5_batch_generate_reports():
    """示例5: 批量生成多只股票报告"""
    print("\n" + "=" * 60)
    print("示例5: 批量生成多只股票报告")
    print("=" * 60)
    
    # 股票列表
    symbols = ['^HSI', '6158.HK', '7200.HK']
    
    visualizer = ExcelVisualizer(output_dir='output/excel')
    analyzer = TechnicalAnalyzer()
    
    print(f"\n正在为 {len(symbols)} 只股票生成报告...")
    
    for symbol in symbols:
        print(f"\n处理 {symbol}...")
        
        try:
            # 获取数据
            data = fetcher.get_historical_data(symbol, period='6mo')
            
            if data is not None and not data.empty:
                # 技术分析
                indicators = analyzer.analyze_stock(symbol, period='6mo')
                
                # 生成报告
                filepath = visualizer.generate_stock_report(
                    symbol=symbol,
                    price_data=data,
                    indicators=indicators
                )
                
                print(f"  ✅ {symbol} 报告已生成")
            else:
                print(f"  ❌ {symbol} 无法获取数据")
        
        except Exception as e:
            print(f"  ❌ {symbol} 错误: {str(e)}")
    
    print(f"\n✅ 批量报告生成完成!")
    print(f"   报告位置: output/excel/")


def example6_launch_web_dashboard():
    """示例6: 启动Web仪表板"""
    print("\n" + "=" * 60)
    print("示例6: 启动Web仪表板")
    print("=" * 60)
    
    print("\n启动方法:")
    print("  在终端运行以下命令:")
    print("\n  cd ai-invest-tool")
    print("  streamlit run src/ai_inv/web_dashboard.py")
    print("\n或在Python中运行:")
    print("\n  import subprocess")
    print("  subprocess.run(['streamlit', 'run', 'src/ai_inv/web_dashboard.py'])")
    
    print("\nWeb界面功能:")
    print("  1. 股票分析 - 实时股票数据和分析")
    print("  2. 技术分析 - 交互式技术指标图表")
    print("  3. AI分析 - ChatGPT驱动的智能分析")
    print("  4. 回测引擎 - 策略回测和结果展示")
    print("  5. 投资组合 - 持仓管理和分析")
    print("  6. 新闻分析 - 新闻情感分析")
    
    print("\n特性:")
    print("  ✅ 实时数据更新")
    print("  ✅ 交互式图表")
    print("  ✅ 响应式设计")
    print("  ✅ Excel导出")
    print("  ✅ 数据可视化")


def example7_combine_with_ai():
    """示例7: 结合AI分析生成完整报告"""
    print("\n" + "=" * 60)
    print("示例7: 结合AI分析生成完整报告")
    print("=" * 60)
    
    print("\n这个示例需要OpenAI API Key")
    print("\n步骤:")
    print("  1. 设置API Key: export OPENAI_API_KEY='your-key'")
    print("  2. 进行技术分析")
    print("  3. 进行AI分析")
    print("  4. 生成包含AI分析的完整报告")
    
    # 示例代码（需要API Key）
    print("\n示例代码:")
    print("""
    from src.ai_inv.ai_analyzer import AIAnalyzer
    from src.ai_inv.excel_visualizer import ExcelVisualizer
    
    # 创建AI分析器
    ai_analyzer = AIAnalyzer()
    
    # 进行AI分析
    ai_result = ai_analyzer.analyze_stock_with_ai(
        symbol='^HSI',
        technical_data={...}
    )
    
    # 生成包含AI分析的报告
    visualizer = ExcelVisualizer()
    filepath = visualizer.generate_stock_report(
        symbol='^HSI',
        price_data=data,
        indicators=indicators,
        ai_analysis=ai_result  # 添加AI分析
    )
    """)


def example8_export_data():
    """示例8: 导出数据到Excel"""
    print("\n" + "=" * 60)
    print("示例8: 导出数据到Excel")
    print("=" * 60)
    
    import pandas as pd
    
    # 创建数据
    print("\n创建示例数据...")
    
    data = {
        '股票代码': ['6158.HK', '7200.HK', '^HSI'],
        '收盘价': [0.52, 10.50, 28500.00],
        '涨跌幅': [2.5, -1.2, 0.8],
        '成交量': [1000000, 500000, 2000000],
        'RSI': [65.5, 45.2, 55.8],
        'MACD': [0.05, -0.02, 12.5],
        '交易信号': ['BUY', 'HOLD', 'BUY'],
        '信号强度': [3, 1, 2]
    }
    
    df = pd.DataFrame(data)
    
    # 导出到Excel
    output_file = 'output/excel/stock_comparison.xlsx'
    os.makedirs(os.path.dirname(output_file), exist_ok=True)
    
    # 使用ExcelWriter
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='股票对比', index=False)
        
        # 添加格式（需要openpyxl）
        try:
            from openpyxl.styles import Font, PatternFill
            from openpyxl.utils import get_column_letter
            
            worksheet = writer.sheets['股票对比']
            
            # 标题行加粗
            for cell in worksheet[1]:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                cell.font = Font(bold=True, color="FFFFFF")
            
            # 自动调整列宽
            for column in worksheet.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 20)
                worksheet.column_dimensions[column_letter].width = adjusted_width
            
            print(f"✅ 数据已导出到: {output_file}")
            print("   并应用了格式化")
        
        except ImportError:
            print("⚠️  openpyxl未安装，仅导出数据（无格式）")
            print("   安装命令: pip install openpyxl")


def main():
    """运行所有示例"""
    print("\n" + "=" * 60)
    print("AI投资工具 - 可视化模块使用示例")
    print("=" * 60)
    
    examples = [
        ("生成股票分析Excel报告", example1_generate_stock_report),
        ("生成回测报告", example2_generate_backtest_report),
        ("生成投资组合报告", example3_generate_portfolio_report),
        ("自定义Excel格式", example4_custom_excel_format),
        ("批量生成多只股票报告", example5_batch_generate_reports),
        ("启动Web仪表板", example6_launch_web_dashboard),
        ("结合AI分析生成完整报告", example7_combine_with_ai),
        ("导出数据到Excel", example8_export_data)
    ]
    
    print("\n可用示例:")
    for i, (name, func) in enumerate(examples, 1):
        print(f"  {i}. {name}")
    
    print("\n运行所有示例...")
    
    # 运行示例1-5和8（需要实际数据的示例）
    example1_generate_stock_report()
    example2_generate_backtest_report()
    example3_generate_portfolio_report()
    example4_custom_excel_format()
    example5_batch_generate_reports()
    example8_export_data()
    
    # 示例6和7是说明性示例
    example6_launch_web_dashboard()
    example7_combine_with_ai()
    
    print("\n" + "=" * 60)
    print("所有示例完成!")
    print("=" * 60)
    
    print("\n提示:")
    print("  - Excel报告保存在 output/excel/ 目录")
    print("  - 使用Excel或WPS打开查看")
    print("  - Web界面使用: streamlit run src/ai_inv/web_dashboard.py")
    print("  - 需要安装依赖: pip install streamlit plotly openpyxl")


if __name__ == "__main__":
    main()
